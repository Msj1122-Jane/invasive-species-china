import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re

SRC = r'D:\OneDrive\Desktop\专题新闻\外来入侵物种图.xlsx'
DST = r'D:\OneDrive\Desktop\专题新闻\外来入侵物种图_带提示词.xlsx'

df = pd.read_excel(SRC)
print(f'Total species: {len(df)}')

# 只处理有明确入侵级别的物种
levels_keep = ['1级','2级','3级','4级','5级','一般','严重','很严重','7级','有待观察']
df_keep = df[df['入侵级别'].isin(levels_keep)].copy()
print(f'Species with level: {len(df_keep)}')

# 科名 → 类别映射
def classify_species(row):
    fn = str(row.get('科名','') or '')
    name = str(row.get('中文名','') or '')
    aliases = str(row.get('别名','') or '')
    origin = str(row.get('原产地','') or '')
    
    # 植物
    plant_families = ['莎草科','禾本科','天南星科','浮萍科','菊科','旋花科','豆科','苋科','茄科',
                      '藜科','马鞭草科','商陆科','大戟科','伞形科','车前科','桔梗科','牻牛儿苗科',
                      '石蒜科','竹芋科','紫葳科','千屈菜科','紫茉莉科','美人蕉科','马兜铃科',
                      '落葵科','锦葵科','玄参科','柳叶菜科','酢浆草科','景天科','兰科','胡颓子科',
                      '荨麻科','爵床科','雨久花科','茜草科','夹竹桃科']
    
    # 动物
    animal_families = ['蚁科','象甲科','天牛科','实蝇科','粉虱科','粉蚧科','蚜科','叶蝉科',
                       '夜蛾科','卷蛾科','螟蛾科','潜叶蛾科','瘿蚊科','果蝇科']
    
    for p in plant_families:
        if p in fn:
            return 'plant'
    for a in animal_families:
        if a in fn:
            return 'insect'
    
    # Heuristic: check common Chinese name patterns
    plant_keywords = ['草','花','莲','树','藻','藤','菊','竹','苔','茅','黍','稗','蓬','葵','兰','菖','芋','萍']
    for kw in plant_keywords:
        if kw in name:
            return 'plant'
    
    insect_keywords = ['蚁','虫','蛾','蜂','蝇','蚁','蚜','虱','螟','蝶','蝽','虻']
    for kw in insect_keywords:
        if kw in name:
            return 'insect'
    
    return 'plant'  # default

def generate_prompt(row):
    name = str(row.get('中文名','') or '').strip()
    family = str(row.get('科名','') or '').strip()
    origin = str(row.get('原产地','') or '').strip()
    aliases = str(row.get('别名','') or '').strip()
    cat = classify_species(row)
    
    # Extract key descriptor from name and aliases
    name_words = name.replace('属','')
    
    # Build detailed description based on name
    if '假高粱' in name or '石茅' in name or name == '石茅':
        desc = '高大禾本科杂草，茎杆粗壮直立，圆锥花序紫色，叶片宽线形，长有根状茎'
    elif '大薸' in name:
        desc = '多年生水生漂浮草本，莲座状叶簇，叶片倒卵状楔形，被白色绒毛，佛焰苞白色'
    elif '凤眼' in name or '水葫芦' in name:
        desc = '水面漂浮草本，膨大叶柄内含气室，蓝紫色花穗，上方花瓣有黄色眼状斑'
    elif '紫茎泽兰' in name or '破坏草' in name:
        desc = '多年生草本，茎紫色被腺毛，对生三角状卵形叶片，白色小花密集成伞房状花序'
    elif '一枝黄花' in name:
        desc = '高大草本，顶生大型圆锥花序，金黄色头状小花密集排列，茎杆直立被柔毛，叶片披针形边缘有锯齿'
    elif '薇甘菊' in name:
        desc = '纤细藤本，心形叶对生，白色小花密集成头状花序，茎节可生根攀附其他植物'
    elif '豚草' in name and '三裂' not in name:
        desc = '一年生草本，羽状深裂叶片，总状雄花序呈碟形下垂，花粉颗粒可见'
    elif '三裂叶豚草' in name:
        desc = '高大一年生草本，三到五裂大型叶片，粗糙被毛，雄花序总状下垂'
    elif '空心' in name or '喜旱' in name:
        desc = '水陆两栖草本，对生长圆形叶片，白色头状花序，匍匐茎节节生根'
    elif '马缨丹' in name or '五色梅' in name:
        desc = '蔓性灌木，茎四棱有倒钩刺，头状花序由黄渐变红，卵形叶对生'
    elif '飞机草' in name:
        desc = '高3-7米亚灌木，对生三角状卵形叶，三出脉明显，淡紫色管状花序排成伞房状'
    elif '水花生' in aliases or name == '喜旱莲子草':
        desc = '匍匐草本，对生倒披针形叶片，白色纸质感头状花序，茎节生不定根'
    elif '鬼针草' in name:
        desc = '一年生草本，羽状复叶三小叶，白色舌状花黄色管状花组成头状花序，黑色条形瘦果具倒钩'
    elif '小蓬草' in name:
        desc = '直立草本，全株绿色被硬毛，线状披针形叶片，小头状花序组成圆锥状'
    elif '藿香蓟' in name:
        desc = '全株被白色柔毛的草本，卵形叶片，蓝色或淡紫色管状花密集成伞房花序'
    elif '毒麦' in name:
        desc = '形似小麦的禾草，茎丛生，穗轴波状曲折，小穗含多个小花，具长芒'
    elif '互花米草' in name:
        desc = '滩涂直立禾草，叶长线形内卷，圆锥花序紧密排列'
    elif '刺苋' in name or '反枝苋' in name:
        desc = '一年生直立草本，茎有棱，卵形叶片有小尖头，圆锥花序顶生，有刺状苞片'
    elif '五爪金龙' in name:
        desc = '多年生缠绕藤本，掌状5深裂叶片，漏斗状粉紫色花冠'
    elif '圆叶牵牛' in name:
        desc = '缠绕藤本，心形全缘叶片被柔毛，漏斗状紫红色花冠'
    elif '银胶菊' in name:
        desc = '一年生草本，二回羽状深裂叶片，头状花序白色'
    elif '加拿大一枝' in name or name == '加拿大一枝黄花':
        desc = '多年生高大草本，披针形互生叶，顶生巨大圆锥花序布满金黄色小花'
    elif '香附子' in name:
        desc = '多年生草本，三棱形茎杆，细长线形叶片基生，红褐色聚伞花序，地下有纺锤形块茎'
    elif '铺地黍' in name:
        desc = '匍匐禾草，节上生根，叶片线状披针形，圆锥花序散开'
    elif '象草' in name:
        desc = '高大丛生禾草，形似甘蔗，粗壮茎杆，长披针形叶片，紫色圆锥花序'
    elif '双穗' in name:
        desc = '匍匐性禾草，节上生根，叶线形，总状花序成对生于秆顶'
    elif '苏门白酒草' in name:
        desc = '全株灰绿色被短糙毛的草本，倒披针形叶片，淡黄或淡紫色头状花序组成大型圆锥状'
    elif '一年蓬' in name:
        desc = '一年生草本，叶长圆形至披针形，白色或淡蓝紫色舌状花环绕黄色管状花'
    elif '假臭草' in name:
        desc = '一年生草本，对生卵形叶有三出脉，蓝紫色头状花序排成伞房状'
    elif '大狼杷草' in name:
        desc = '一年生草本，羽状复叶3-5小叶，黄色头状花序，瘦果顶端具2芒刺'
    elif '三裂叶薯' in name:
        desc = '缠绕藤本，叶片三裂或心形，漏斗状粉红色花冠'
    elif '光荚含羞草' in name:
        desc = '落叶灌木，二回羽状复叶触之闭合，白色头状花序球形，茎有疏刺'
    elif '落葵薯' in name:
        desc = '多年生缠绕藤本，肉质心形叶片，穗状花序下垂，白色小花香气浓郁'
    elif '土荆芥' in name:
        desc = '一年生草本有强烈气味，长圆形叶片边缘波状，穗状花序腋生，花小黄绿色'
    elif '肿柄菊' in name:
        desc = '高大草本或亚灌木，大型掌状叶片，头状花序金黄色，花序梗顶部膨大'
    else:
        # Generic template by category
        if cat == 'plant':
            if family:
                desc = f'{family}植物，{name}植株形态特写'
            else:
                desc = f'{name}植株形态特写'
        else:
            desc = f'{name}成虫形态特写'
    
    # Add origin/environment context
    env = '自然光照'
    if origin and str(origin).lower() not in ['不详','nan','']:
        env = str(origin) + '原生境'
    
    # Add family info if not already included
    family_part = ''
    if family and str(family).lower() not in ['nan',''] and str(family) not in desc:
        family_part = f'（{str(family)}）'
    
    # Clean NaN artifacts
    desc = desc.replace('nan植物，','').replace('nan','')
    
    prompt = f'高清植物科学图鉴摄影，{desc}，白色背景，{env}，自然光，细节清晰，分类学展示照{family_part}'
    prompt = prompt.replace('（nan）','').replace('  ',' ').replace('，，', '，').strip()
    return prompt

df_keep['AI提示词'] = df_keep.apply(generate_prompt, axis=1)

# 用 openpyxl 写入，保留原文件格式
wb = load_workbook(SRC)
ws = wb.active

# Find header row (first row)
header_row = 1
max_col = ws.max_column

# Check if 'AI提示词' column already exists
ai_col = None
for col in range(1, max_col + 1):
    if ws.cell(row=header_row, column=col).value == 'AI提示词':
        ai_col = col
        break

if ai_col is None:
    ai_col = max_col + 1
    ws.cell(row=header_row, column=ai_col, value='AI提示词')
    src_font = ws.cell(row=header_row, column=1).font
    ws.cell(row=header_row, column=ai_col).font = src_font.copy() if src_font else None
    # Adjust column width
    ws.column_dimensions[get_column_letter(ai_col)].width = 80

# Write prompts for kept species
for idx, prompt in enumerate(df_keep['AI提示词']):
    orig_idx = df_keep.index[idx]
    excel_row = orig_idx + 2  # +2 for 0-index and header
    cell = ws.cell(row=excel_row, column=ai_col)
    cell.value = prompt

wb.save(DST)
print(f'Written {len(df_keep)} prompts to {DST}, column {get_column_letter(ai_col)}')
print('Done.')
