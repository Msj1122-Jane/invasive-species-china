#!/usr/bin/env python3
"""
从 iflora_外来入侵物种.xlsx 提取数据，生成：
1. species_341.json — 341种外来入侵物种完整数据
2. species_province_matrix.json — 329种 × 35省 二元矩阵
"""
import json
import openpyxl
import os

SRC = r'D:\OneDrive\Desktop\专题新闻\iflora_外来入侵物种.xlsx'
OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')

wb = openpyxl.load_workbook(SRC, data_only=True)

# ============================================================
# 1. species_341.json — 来自"省份分布" sheet
# ============================================================
ws = wb['省份分布']
headers = ['species_id', 'name_cn', 'name_latin', 'category', 'family', 'genus',
           'level', 'origin', 'provinces_raw', 'province_count', 'harm_description', 'control_method']

species_list = []
for r in range(2, ws.max_row + 1):
    sp = {}
    sp['species_id'] = ws.cell(row=r, column=1).value
    sp['name_cn'] = (ws.cell(row=r, column=2).value or '').strip()
    sp['name_latin'] = (ws.cell(row=r, column=3).value or '').strip()
    sp['category'] = (ws.cell(row=r, column=4).value or '未知').strip()
    sp['family'] = (ws.cell(row=r, column=5).value or '未知').strip()
    sp['genus'] = (ws.cell(row=r, column=6).value or '未知').strip()

    # 清洗入侵级别为数字
    raw_level = str(ws.cell(row=r, column=7).value or '').strip()
    level_num = None
    for candidate in ['1', '2', '3', '4', '5']:
        if candidate in raw_level:
            level_num = int(candidate)
            break
    if level_num is None:
        # fallback mappings
        level_map = {'恶性': 1, '很严重': 2, '严重': 2, '局部': 3, '一般': 4, '有待观察': 5}
        for k, v in level_map.items():
            if k in raw_level:
                level_num = v
                break
    sp['level'] = level_num if level_num else 5

    sp['origin'] = (ws.cell(row=r, column=8).value or '未知').strip()

    # 省市列表清洗
    prov_raw = (ws.cell(row=r, column=9).value or '').strip()
    prov_list = [p.strip() for p in prov_raw.replace('、', ',').split(',') if p.strip()]
    # 统一省市名称
    province_standard = {
        '澳门': '中国澳门', '香港': '中国香港', '台湾': '中国台湾',
        '南海': '海南', '南海诸岛': '海南'
    }
    prov_list = [province_standard.get(p, p) for p in prov_list]
    # Deduplicate after merging (南海诸岛 → 海南)
    prov_list = list(dict.fromkeys(prov_list))
    sp['province_list'] = prov_list
    sp['province_count'] = int(ws.cell(row=r, column=10).value or len(prov_list))

    sp['harm_description'] = (ws.cell(row=r, column=11).value or '').strip()
    sp['control_method'] = (ws.cell(row=r, column=12).value or '').strip()

    # 大类归类 (用于旭日图等)
    if sp['category'] and '被子' in sp['category']:
        sp['major_group'] = '被子植物'
    elif sp['category'] and '蕨类' in sp['category']:
        sp['major_group'] = '蕨类植物'
    elif sp['category'] and '裸子' in sp['category']:
        sp['major_group'] = '裸子植物'
    elif sp['category'] and '苔藓' in sp['category']:
        sp['major_group'] = '苔藓植物'
    else:
        sp['major_group'] = sp['category'] if sp['category'] != '未知' else '其他'

    # 原产地大洲归类
    origin_map = {
        '美洲': '美洲', '北美洲': '美洲', '南美洲': '美洲', '中美洲': '美洲', '热带美洲': '美洲',
        '欧洲': '欧洲', '地中海': '欧洲', '地中海地区': '欧洲', '地中海沿岸': '欧洲',
        '非洲': '非洲', '热带非洲': '非洲', '南非': '非洲',
        '亚洲': '亚洲', '热带亚洲': '亚洲', '东南亚': '亚洲',
        '大洋洲': '大洋洲', '澳大利亚': '大洋洲',
        '墨西哥': '美洲', '巴西': '美洲', '印度': '亚洲', '中国': '亚洲',
        '日本': '亚洲', '秘鲁': '美洲', '阿根廷': '美洲', '智利': '美洲',
        '伊朗': '亚洲', '伊拉克': '亚洲', '土耳其': '亚洲',
        '埃塞俄比亚': '非洲', '马达加斯加': '非洲', '肯尼亚': '非洲',
    }
    sp['origin_continent'] = origin_map.get(sp['origin'], '其他')

    species_list.append(sp)

# 按级别排序
species_list.sort(key=lambda x: (x['level'], x['name_cn']))

with open(os.path.join(OUT_DIR, 'species_341.json'), 'w', encoding='utf-8') as f:
    json.dump(species_list, f, ensure_ascii=False, indent=2)

print(f'[OK] species_341.json: {len(species_list)} species')
levels_dist = {}
for sp in species_list:
    lv = sp['level']
    levels_dist[lv] = levels_dist.get(lv, 0) + 1
print(f'  级别分布: {dict(sorted(levels_dist.items()))}')

# ============================================================
# 2. species_province_matrix.json — 来自"物种×省份" sheet
# ============================================================
ws2 = wb['物种×省份']

# Build matrix
species_province = {}  # {species_name: {province1: level, province2: level, ...}}
species_info = {}  # {species_name: {name_latin, level}}
all_provinces = set()
all_species = set()

for r in range(2, ws2.max_row + 1):
    name_cn = (ws2.cell(row=r, column=1).value or '').strip()
    name_latin = (ws2.cell(row=r, column=2).value or '').strip()
    level_raw = str(ws2.cell(row=r, column=3).value or '').strip()
    province = (ws2.cell(row=r, column=4).value or '').strip()

    if not name_cn or not province:
        continue

    # 统一省份
    province_standard = {
        '澳门': '中国澳门', '香港': '中国香港', '台湾': '中国台湾',
        '南海': '海南', '南海诸岛': '海南'
    }
    province = province_standard.get(province, province)

    # 清洗级别
    level_num = 5
    for candidate in ['1', '2', '3', '4', '5']:
        if candidate in level_raw:
            level_num = int(candidate)
            break

    if name_cn not in species_province:
        species_province[name_cn] = {}
    species_province[name_cn][province] = level_num
    species_info[name_cn] = {'name_latin': name_latin, 'level': level_num}
    all_provinces.add(province)
    all_species.add(name_cn)

# 构建矩阵结构
province_list_sorted = sorted(all_provinces)
species_list_sorted = sorted(all_species, key=lambda s: (species_info[s]['level'], s))

# Binary matrix rows
matrix_rows = []
for name_cn in species_list_sorted:
    info = species_info[name_cn]
    prov_dict = species_province[name_cn]
    row = {
        'name_cn': name_cn,
        'name_latin': info['name_latin'],
        'level': info['level'],
        'province_count': len(prov_dict),
        'provinces': {p: prov_dict.get(p, 0) for p in province_list_sorted}  # 0=absent, 1-5=level
    }
    matrix_rows.append(row)

# 省份汇总
province_summary = []
for p in province_list_sorted:
    sp_in_prov = [name_cn for name_cn in species_list_sorted
                  if species_province.get(name_cn, {}).get(p, 0) > 0]
    levels_count = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for name_cn in sp_in_prov:
        lv = species_province[name_cn][p]
        levels_count[lv] = levels_count.get(lv, 0) + 1
    province_summary.append({
        'name': p,
        'total': len(sp_in_prov),
        'level_1': levels_count[1],
        'level_2': levels_count[2],
        'level_3': levels_count[3],
        'level_4': levels_count[4],
        'level_5': levels_count[5],
    })

output = {
    'provinces': province_list_sorted,
    'province_summary': province_summary,
    'species': matrix_rows,
    'total_species': len(matrix_rows),
    'total_provinces': len(province_list_sorted),
    'total_links': sum(len(species_province[s]) for s in species_province)
}

with open(os.path.join(OUT_DIR, 'species_province_matrix.json'), 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f'[OK] species_province_matrix.json: {len(matrix_rows)} species × {len(province_list_sorted)} provinces, {output["total_links"]} links')

# ============================================================
# 3. 额外：省份统计JSON（用于省份热力图）
# ============================================================
ws3 = wb['省份统计']
prov_stats = []
for r in range(2, ws3.max_row + 1):
    prov = (ws3.cell(row=r, column=1).value or '').strip()
    # 统一省份
    province_standard = {
        '澳门': '中国澳门', '香港': '中国香港', '台湾': '中国台湾',
        '南海': '海南', '南海诸岛': '海南'
    }
    prov = province_standard.get(prov, prov)
    # Deduplicate: if another row mapped to same province, merge totals
    existing = next((ps for ps in prov_stats if ps['name'] == prov), None)
    if existing:
        existing['total'] += int(ws3.cell(row=r, column=2).value or 0)
        existing['level_1'] += int(ws3.cell(row=r, column=3).value or 0)
        existing['level_2'] += int(ws3.cell(row=r, column=4).value or 0)
        existing['level_3'] += int(ws3.cell(row=r, column=5).value or 0)
        existing['level_4'] += int(ws3.cell(row=r, column=6).value or 0)
        existing['level_5'] += int(ws3.cell(row=r, column=7).value or 0)
        continue
    prov_stats.append({
        'name': prov,
        'total': int(ws3.cell(row=r, column=2).value or 0),
        'level_1': int(ws3.cell(row=r, column=3).value or 0),
        'level_2': int(ws3.cell(row=r, column=4).value or 0),
        'level_3': int(ws3.cell(row=r, column=5).value or 0),
        'level_4': int(ws3.cell(row=r, column=6).value or 0),
        'level_5': int(ws3.cell(row=r, column=7).value or 0),
    })

with open(os.path.join(OUT_DIR, 'province_stats_iflora.json'), 'w', encoding='utf-8') as f:
    json.dump(prov_stats, f, ensure_ascii=False, indent=2)

print(f'[OK] province_stats_iflora.json: {len(prov_stats)} provinces')
top5 = sorted(prov_stats, key=lambda x: -x['total'])[:5]
for p in top5:
    print(f'  {p["name"]}: {p["total"]} species (L1={p["level_1"]}, L2={p["level_2"]}, L3={p["level_3"]}, L4={p["level_4"]}, L5={p["level_5"]})')

# ============================================================
# 4. 级别汇总（用于入侵金字塔）
# ============================================================
ws4 = wb['级别汇总']
level_summary = []
pyramid_data = {}
for r in range(2, ws4.max_row + 1):
    level = str(ws4.cell(row=r, column=1).value or '').strip()
    count = int(ws4.cell(row=r, column=2).value or 0)
    desc = (ws4.cell(row=r, column=3).value or '').strip()
    level_summary.append({'level': level, 'count': count, 'description': desc})
    pyramid_data[level] = {'count': count, 'description': desc}

with open(os.path.join(OUT_DIR, 'level_summary_iflora.json'), 'w', encoding='utf-8') as f:
    json.dump({'pyramid': pyramid_data, 'levels': level_summary}, f, ensure_ascii=False, indent=2)

print(f'[OK] level_summary_iflora.json: {len(level_summary)} categories')
print('  Pyramid:', {k: v['count'] for k, v in pyramid_data.items()})

print('\n=== 全部完成 ===')
